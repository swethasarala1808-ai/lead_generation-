import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, time

NAVY  = "0F1B2D"
SILVER= "E2E8F0"
ALT   = "F0F4FF"
ACCENT= "1967D2"
GOLD  = "F5A623"
GREEN = "00B050"

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

class ExcelExporter:

    def export(self, leads: list, job_id: str, source: str = "") -> str:
        if not leads:
            raise ValueError("No leads to export")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"
        border = _thin_border()
        headers = list(leads[0].keys())

        # ── Title ──
        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        tc = ws["A1"]
        tc.value = "🎯  LeadGen Pro — Quality B2B Lead Report"
        tc.font  = Font(name="Calibri", bold=True, size=16, color=GOLD)
        tc.fill  = PatternFill("solid", fgColor=NAVY)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36

        ws.merge_cells(f"A2:{get_column_letter(len(headers))}2")
        sub = ws["A2"]
        src_label = "IndiaMART (Live)" if "indiamart" in source else ("Google Places (Live)" if "google" in source else "Mixed Sources")
        sub.value = f"Source: {src_label}  |  Leads: {len(leads)}  |  Exported: {time.strftime('%d %b %Y %H:%M')}"
        sub.font  = Font(name="Calibri", italic=True, size=10, color="94A3B8")
        sub.fill  = PatternFill("solid", fgColor=NAVY)
        sub.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        # ── Header row ──
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.font      = Font(name="Calibri", bold=True, size=10, color=SILVER)
            cell.fill      = PatternFill("solid", fgColor=ACCENT)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = border
        ws.row_dimensions[3].height = 30

        # ── Data ──
        for ri, lead in enumerate(leads, 4):
            is_alt = (ri % 2 == 0)
            for ci, h in enumerate(headers, 1):
                val  = lead.get(h, "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font      = Font(name="Calibri", size=9)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border    = border
                if is_alt:
                    cell.fill = PatternFill("solid", fgColor=ALT)
                # colour accents
                if h == "Lead Quality Score" and val:
                    n = int(str(val).replace("%",""))
                    if n >= 85:
                        cell.font = Font(name="Calibri", size=9, bold=True, color=GREEN)
                    elif n >= 70:
                        cell.font = Font(name="Calibri", size=9, color="FFA500")
                if h == "Source" and "IndiaMART" in str(val):
                    cell.font = Font(name="Calibri", size=9, bold=True, color="C00000")
                if h == "Source" and "Google" in str(val):
                    cell.font = Font(name="Calibri", size=9, bold=True, color=ACCENT)
                if h in ("Mobile Number","Buyer Mobile") and val:
                    cell.font = Font(name="Calibri", size=9, color="047857")

        # ── Column widths ──
        WIDTHS = {
            "Sr No":8,"Company Name":28,"Buyer Name":22,"Business Type":20,
            "Registration Type":20,"Owner / Contact Person":22,"Mobile Number":16,
            "Buyer Mobile":16,"Alternate Mobile":16,"Email ID":28,"Buyer Email":28,
            "Website":26,"Address":32,"Buyer Address":32,"City":14,"Buyer City":14,
            "State":16,"Buyer State":16,"Pincode":10,"Google Rating":14,
            "No. of Reviews":14,"Opening Hours":26,"Lead Quality Score":16,
            "WhatsApp Available":14,"Google Maps Link":28,"Source":22,
            "Remarks":24,"Product / Requirement":28,"Buyer Message":40,
            "Lead Type":20,"Received At":20,"Lead ID":14,
        }
        for ci, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(ci)].width = WIDTHS.get(h, 16)

        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{len(leads)+3}"

        # ── Summary sheet ──
        ws2 = wb.create_sheet("Summary")
        ws2.merge_cells("A1:C1")
        c = ws2["A1"]
        c.value = "Lead Generation Summary"
        c.font  = Font(name="Calibri", bold=True, size=14, color=SILVER)
        c.fill  = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
        rows = [
            ("Total Leads",     len(leads)),
            ("Source",          src_label),
            ("Export Date",     time.strftime("%d %b %Y %H:%M")),
            ("With Phone",      sum(1 for l in leads if l.get("Mobile Number") or l.get("Buyer Mobile"))),
            ("With Email",      sum(1 for l in leads if l.get("Email ID") or l.get("Buyer Email"))),
            ("With Website",    sum(1 for l in leads if l.get("Website") or l.get("Full Website URL"))),
        ]
        if "google" in source:
            rows += [
                ("High Rated (≥4★)", sum(1 for l in leads if l.get("Google Rating","").startswith(("4","5")))),
                ("High Quality >85%", sum(1 for l in leads if str(l.get("Lead Quality Score","")).replace("%","").isdigit() and int(str(l.get("Lead Quality Score","0%")).replace("%","")) >= 85)),
            ]
        if "indiamart" in source:
            rows += [
                ("Web Enquiries",   sum(1 for l in leads if "Web" in str(l.get("Lead Type","")))),
                ("BuyLeads",        sum(1 for l in leads if "Buy" in str(l.get("Lead Type","")))),
                ("PNS Calls",       sum(1 for l in leads if "PNS" in str(l.get("Lead Type","")))),
            ]
        for ri, (lbl, val) in enumerate(rows, 3):
            ws2.cell(row=ri, column=1, value=lbl).font = Font(name="Calibri", bold=True, size=10)
            ws2.cell(row=ri, column=2, value=val).font = Font(name="Calibri", size=10)
            ws2.cell(row=ri, column=1).fill = PatternFill("solid", fgColor="E8EAF6")
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 22

        os.makedirs("/home/claude/lead_generation-/exports", exist_ok=True)
        path = f"/home/claude/lead_generation-/exports/leads_{job_id}.xlsx"
        wb.save(path)
        return path
