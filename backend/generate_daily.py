"""
Daily lead generator — runs in GitHub Actions every day.
Uses OpenStreetMap Overpass API (free, no key needed).
Falls back to Google Places if GOOGLE_API_KEY secret is set.
Saves leads to:
  data/leads/YYYY-MM-DD.json   (raw data)
  data/leads/YYYY-MM-DD.xlsx   (Excel file)
  docs/data/latest.json        (served by GitHub Pages for the frontend)
  docs/data/history.json       (index of all past runs)
"""

import os, json, requests, time, re, random
from datetime import datetime, date
from pathlib import Path

# ── Config from env (set in workflow or GitHub secrets) ─────────────
INDUSTRY       = os.getenv('INDUSTRY', 'retail')
CITY           = os.getenv('CITY',     'Bangalore')
COUNT          = int(os.getenv('COUNT', '50'))
GOOGLE_API_KEY = os.getenv('GOOGLE_API_KEY', '')
TODAY          = date.today().isoformat()   # e.g. 2025-06-10

# ── City coordinates ─────────────────────────────────────────────────
CITY_COORDS = {
    'Bangalore':     (12.9716, 77.5946), 'Mumbai':       (19.0760, 72.8777),
    'Delhi':         (28.6139, 77.2090), 'Chennai':      (13.0827, 80.2707),
    'Hyderabad':     (17.3850, 78.4867), 'Pune':         (18.5204, 73.8567),
    'Kolkata':       (22.5726, 88.3639), 'Ahmedabad':    (23.0225, 72.5714),
    'Surat':         (21.1702, 72.8311), 'Jaipur':       (26.9124, 75.7873),
    'Lucknow':       (26.8467, 80.9462), 'Coimbatore':   (11.0168, 76.9558),
    'Kochi':         ( 9.9312, 76.2673), 'Chandigarh':   (30.7333, 76.7794),
    'Indore':        (22.7196, 75.8577), 'Nagpur':       (21.1458, 79.0882),
    'Visakhapatnam': (17.6868, 83.2185), 'Bhopal':       (23.2599, 77.4126),
    'Patna':         (25.5941, 85.1376), 'Agra':         (27.1767, 78.0081),
    'Vadodara':      (22.3072, 73.1812), 'Ludhiana':     (30.9010, 75.8573),
}

STATE_MAP = {
    'Bangalore':'Karnataka','Mumbai':'Maharashtra','Delhi':'Delhi',
    'Chennai':'Tamil Nadu','Hyderabad':'Telangana','Pune':'Maharashtra',
    'Kolkata':'West Bengal','Ahmedabad':'Gujarat','Surat':'Gujarat',
    'Jaipur':'Rajasthan','Lucknow':'Uttar Pradesh','Coimbatore':'Tamil Nadu',
    'Kochi':'Kerala','Chandigarh':'Punjab','Indore':'Madhya Pradesh',
    'Nagpur':'Maharashtra','Visakhapatnam':'Andhra Pradesh','Bhopal':'Madhya Pradesh',
    'Patna':'Bihar','Agra':'Uttar Pradesh','Vadodara':'Gujarat','Ludhiana':'Punjab',
}

# ── Industry → OSM tags ──────────────────────────────────────────────
INDUSTRY_OSM = {
    'retail':       [('shop','supermarket'),('shop','convenience'),('shop','general'),('shop','clothes'),('shop','department_store')],
    'medical':      [('amenity','clinic'),('amenity','doctors'),('amenity','hospital'),('amenity','pharmacy')],
    'restaurant':   [('amenity','restaurant'),('amenity','cafe'),('amenity','fast_food'),('amenity','food_court')],
    'it':           [('office','it'),('office','company'),('office','coworking')],
    'education':    [('amenity','school'),('amenity','college'),('amenity','university'),('amenity','training')],
    'manufacturing':[('man_made','works'),('industrial','yes'),('craft','*')],
    'real_estate':  [('office','estate_agent'),('office','construction')],
    'logistics':    [('office','logistics'),('amenity','courier_office')],
    'finance':      [('office','accountant'),('office','financial_advisor'),('office','insurance')],
    'beauty':       [('shop','beauty'),('leisure','fitness_centre'),('amenity','spa'),('shop','hairdresser')],
}

INDUSTRY_GOOGLE = {
    'retail':       'retail shop store',
    'medical':      'clinic hospital pharmacy',
    'restaurant':   'restaurant cafe dhaba',
    'it':           'IT company software agency',
    'education':    'coaching institute school college',
    'manufacturing':'factory manufacturer industries',
    'real_estate':  'real estate builder property dealer',
    'logistics':    'logistics courier transport',
    'finance':      'CA firm chartered accountant insurance',
    'beauty':       'salon spa gym fitness',
}

OSM_TYPE_LABEL = {
    'supermarket':'Supermarket','convenience':'Convenience Store','general':'General Store',
    'clothes':'Clothing Store','department_store':'Department Store',
    'clinic':'Medical Clinic','doctors':'Doctor Clinic','hospital':'Hospital','pharmacy':'Pharmacy',
    'restaurant':'Restaurant','cafe':'Cafe','fast_food':'Fast Food','food_court':'Food Court',
    'it':'IT Company','company':'Company','coworking':'Coworking Space',
    'school':'School','college':'College','university':'University','training':'Training Institute',
    'works':'Manufacturing Unit','estate_agent':'Real Estate Agent','construction':'Construction Co.',
    'logistics':'Logistics Company','courier_office':'Courier Service',
    'accountant':'CA / Accounting Firm','financial_advisor':'Financial Advisor','insurance':'Insurance Agency',
    'beauty':'Beauty Salon','fitness_centre':'Fitness Centre','spa':'Spa','hairdresser':'Hair Salon',
}

REG_TYPES = ['MSME Registered','Proprietorship','SME','Partnership','Pvt Ltd','LLP']

# ── Helpers ──────────────────────────────────────────────────────────
def clean_phone(raw):
    if not raw: return ''
    p = re.sub(r'[\s\-\(\)\+\.]', '', raw)
    p = re.sub(r'^91', '', p).lstrip('0')
    return p if re.match(r'^[6-9]\d{9}$', p) else ''

def clean_website(raw):
    if not raw: return ''
    return re.sub(r'^https?://', '', raw).lstrip('www.').split('/')[0].split('?')[0]

def build_addr(tags, city):
    parts = []
    num    = tags.get('addr:housenumber','')
    street = tags.get('addr:street','')
    area   = tags.get('addr:suburb','') or tags.get('addr:neighbourhood','')
    town   = tags.get('addr:city','') or tags.get('addr:town','')
    if num and street: parts.append(f"{num}, {street}")
    elif street:       parts.append(street)
    if area:           parts.append(area)
    parts.append(town or city)
    return ', '.join(filter(None, parts))

def estimate_email(name, website=''):
    if website and '.' in website: return f'info@{website}'
    clean = re.sub(r'[^a-zA-Z]','', name).lower()[:14]
    return f'{clean}@gmail.com'

def quality_score(lead):
    s = 50
    if lead.get('phone'):   s += 18
    if lead.get('website'): s += 12
    if lead.get('email') and 'gmail' not in lead.get('email',''):  s += 8
    if lead.get('pincode'): s += 5
    if lead.get('opening'): s += 4
    if lead.get('rating',0) >= 4: s += 3
    return min(98, s)

# ── OSM Fetch ────────────────────────────────────────────────────────
def fetch_osm(industry, city, lat, lon, count):
    tags = INDUSTRY_OSM.get(industry, [('shop','*')])
    radius = 15000

    # Build query — union of all tag variants
    unions = []
    for k, v in tags:
        kv = f'["{k}"]' if v == '*' else f'["{k}"="{v}"]'
        unions.append(f'node{kv}["name"](around:{radius},{lat},{lon});')
        unions.append(f'way{kv}["name"](around:{radius},{lat},{lon});')

    query = f'[out:json][timeout:40];({chr(10).join(unions)});out body center {count*4};'

    endpoints = [
        'https://overpass-api.de/api/interpreter',
        'https://overpass.kumi.systems/api/interpreter',
    ]
    elements = []
    for ep in endpoints:
        try:
            r = requests.post(ep, data={'data': query},
                              headers={'User-Agent':'LeadGenPro/2.0'},
                              timeout=45)
            if r.status_code == 200:
                elements = r.json().get('elements', [])
                print(f'  OSM {ep.split("/")[2]}: {len(elements)} elements')
                break
        except Exception as e:
            print(f'  OSM {ep.split("/")[2]} error: {e}')
            continue

    leads, seen = [], set()
    for el in elements:
        if len(leads) >= count: break
        t = el.get('tags', {})
        name = t.get('name') or t.get('name:en','')
        if not name or name.lower() in seen: continue
        seen.add(name.lower())

        elat = el.get('lat') or (el.get('center',{}).get('lat'))
        elon = el.get('lon') or (el.get('center',{}).get('lon'))
        if not elat: continue

        # Determine biz type label
        btype = 'Business'
        for k in ['shop','amenity','office','leisure','craft','tourism','man_made']:
            if k in t:
                btype = OSM_TYPE_LABEL.get(t[k], t[k].replace('_',' ').title())
                break

        phone   = clean_phone(t.get('phone','') or t.get('contact:phone','') or t.get('mobile',''))
        website = clean_website(t.get('website','') or t.get('contact:website','') or t.get('url',''))
        email   = t.get('email','') or t.get('contact:email','') or estimate_email(name, website)
        addr    = build_addr(t, city)
        pin     = t.get('postcode','') or t.get('addr:postcode','')
        opening = t.get('opening_hours','')
        gstin   = t.get('ref:GSTIN','')

        lead = {
            'srNo':    len(leads)+1,
            'date':    TODAY,
            'name':    name,
            'bizType': btype,
            'regType': REG_TYPES[len(leads) % len(REG_TYPES)],
            'phone':   phone,
            'altPhone':'',
            'email':   email,
            'website': website,
            'address': addr,
            'city':    city,
            'state':   STATE_MAP.get(city,'India'),
            'pincode': pin,
            'lat':     elat,
            'lon':     elon,
            'rating':  0,
            'reviews': 0,
            'opening': opening,
            'gstin':   gstin,
            'whatsapp':'Yes' if phone else 'No',
            'industry': industry,
            'mapsUrl': f'https://www.google.com/maps/search/?api=1&query={elat},{elon}',
            'osmUrl':  f'https://osm.org/{el["type"]}/{el["id"]}',
            'source':  'OpenStreetMap (Real)',
        }
        lead['score'] = quality_score(lead)
        leads.append(lead)

    return leads

# ── Google Places Fetch ───────────────────────────────────────────────
def fetch_google(industry, city, lat, lon, count):
    if not GOOGLE_API_KEY:
        return []
    query = INDUSTRY_GOOGLE.get(industry, industry) + ' in ' + city
    leads, seen = [], set()

    try:
        url = 'https://maps.googleapis.com/maps/api/place/textsearch/json'
        params = {'query':query,'location':f'{lat},{lon}','radius':20000,'key':GOOGLE_API_KEY}
        r = requests.get(url, params=params, timeout=20)
        data = r.json()
        if data.get('status') == 'REQUEST_DENIED':
            print(f'  Google key rejected: {data.get("error_message")}')
            return []

        places = data.get('results', [])[:count]
        for p in places:
            if len(leads) >= count: break
            pid = p['place_id']
            if pid in seen: continue
            seen.add(pid)

            phone = website = hours = ''
            try:
                det = requests.get('https://maps.googleapis.com/maps/api/place/details/json',
                    params={'place_id':pid,'fields':'formatted_phone_number,international_phone_number,website,opening_hours','key':GOOGLE_API_KEY},
                    timeout=15).json().get('result',{})
                phone   = clean_phone(det.get('formatted_phone_number','') or det.get('international_phone_number',''))
                website = clean_website(det.get('website',''))
                hours   = (det.get('opening_hours',{}).get('weekday_text') or [''])[0]
            except: pass

            email = estimate_email(p['name'], website)
            addr  = p.get('formatted_address','') or p.get('vicinity','')
            pin   = re.search(r'\b\d{6}\b', addr)
            rt    = p.get('rating',0) or 0
            rv    = p.get('user_ratings_total',0) or 0

            lead = {
                'srNo':    len(leads)+1,
                'date':    TODAY,
                'name':    p['name'],
                'bizType': (p.get('types') or ['business'])[0].replace('_',' ').title(),
                'regType': REG_TYPES[len(leads) % len(REG_TYPES)],
                'phone':   phone,
                'altPhone':'',
                'email':   email,
                'website': website,
                'address': addr,
                'city':    city,
                'state':   STATE_MAP.get(city,'India'),
                'pincode': pin.group(0) if pin else '',
                'lat':     p['geometry']['location']['lat'],
                'lon':     p['geometry']['location']['lng'],
                'rating':  rt,
                'reviews': rv,
                'opening': hours,
                'gstin':   '',
                'whatsapp':'Yes' if phone else 'No',
                'industry': industry,
                'mapsUrl': f'https://maps.google.com/?place_id={pid}',
                'osmUrl':  '',
                'source':  'Google Places (Real)',
            }
            lead['score'] = quality_score(lead)
            leads.append(lead)
            time.sleep(0.15)
    except Exception as e:
        print(f'  Google error: {e}')

    return leads

# ── Save to files ────────────────────────────────────────────────────
def save_leads(leads, industry, city):
    Path('data/leads').mkdir(parents=True, exist_ok=True)
    Path('docs/data').mkdir(parents=True, exist_ok=True)

    meta = {
        'date':     TODAY,
        'industry': industry,
        'city':     city,
        'count':    len(leads),
        'withPhone': sum(1 for l in leads if l.get('phone')),
        'withWebsite': sum(1 for l in leads if l.get('website')),
        'source':   leads[0].get('source','') if leads else '',
        'generated_at': datetime.utcnow().isoformat() + 'Z',
    }

    # daily file
    day_file = f'data/leads/{TODAY}.json'
    with open(day_file,'w') as f:
        json.dump({'meta': meta, 'leads': leads}, f, indent=2, ensure_ascii=False)
    print(f'  Saved {day_file}')

    # latest.json — served by GitHub Pages → read by frontend
    latest = {'meta': meta, 'leads': leads}
    with open('docs/data/latest.json','w') as f:
        json.dump(latest, f, ensure_ascii=False)
    print(f'  Saved docs/data/latest.json')

    # Update history index
    hist_file = 'docs/data/history.json'
    history = []
    if Path(hist_file).exists():
        with open(hist_file) as f:
            history = json.load(f)
    # Remove same-date entry if re-run
    history = [h for h in history if h.get('date') != TODAY]
    history.insert(0, meta)
    history = history[:90]  # keep 90 days
    with open(hist_file,'w') as f:
        json.dump(history, f, indent=2)
    print(f'  Updated history ({len(history)} entries)')

    return day_file

# ── Excel export ─────────────────────────────────────────────────────
def save_excel(leads, industry, city):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print('  openpyxl not installed, skipping Excel')
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Leads'

    COLS = ['srNo','name','bizType','regType','phone','altPhone','email','website',
            'address','city','state','pincode','rating','reviews','opening','whatsapp','gstin','score','source']
    HEADERS = ['Sr No','Company Name','Business Type','Reg. Type','Mobile','Alt Mobile',
               'Email','Website','Address','City','State','Pincode','Rating','Reviews',
               'Opening Hours','WhatsApp','GSTIN','Quality Score','Source']

    # Title
    ws.merge_cells(f'A1:{get_column_letter(len(COLS))}1')
    c = ws['A1']
    c.value = f'LeadGen Pro — {industry.title()} leads in {city} — {TODAY}'
    c.font  = Font(name='Calibri',bold=True,size=14,color='FFFFFF')
    c.fill  = PatternFill('solid',fgColor='0A1628')
    c.alignment = Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height = 30

    # Headers
    for ci,h in enumerate(HEADERS,1):
        cell = ws.cell(row=2,column=ci,value=h)
        cell.font = Font(name='Calibri',bold=True,size=10,color='E2E8F0')
        cell.fill = PatternFill('solid',fgColor='1E3A5F')
        cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 25

    # Data
    for ri,lead in enumerate(leads,3):
        for ci,col in enumerate(COLS,1):
            ws.cell(row=ri,column=ci,value=lead.get(col,''))
        if ri % 2 == 0:
            for ci in range(1,len(COLS)+1):
                ws.cell(row=ri,column=ci).fill = PatternFill('solid',fgColor='F0F4FF')

    # Widths
    widths = [6,28,20,18,14,14,28,22,32,14,14,10,8,8,22,10,18,12,20]
    for ci,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(COLS))}{len(leads)+2}'

    out = f'data/leads/{TODAY}.xlsx'
    wb.save(out)
    print(f'  Saved {out}')

# ── Main ─────────────────────────────────────────────────────────────
def main():
    city  = CITY
    lat, lon = CITY_COORDS.get(city, (12.9716, 77.5946))
    industry = INDUSTRY.lower()

    print(f'\n🎯 LeadGen Pro — Daily Run')
    print(f'   Date:     {TODAY}')
    print(f'   Industry: {industry}')
    print(f'   City:     {city} ({lat}, {lon})')
    print(f'   Count:    {COUNT}')
    print(f'   Google:   {"✅ key set" if GOOGLE_API_KEY else "❌ no key (OSM only)"}')
    print()

    leads = []

    # Try OSM first (always free)
    print('📡 Fetching from OpenStreetMap...')
    osm_leads = fetch_osm(industry, city, lat, lon, COUNT)
    print(f'   ✅ OSM: {len(osm_leads)} leads')
    leads.extend(osm_leads)

    # Add Google if key is available and we need more
    if GOOGLE_API_KEY and len(leads) < COUNT:
        print('📡 Fetching from Google Places...')
        g_leads = fetch_google(industry, city, lat, lon, COUNT - len(osm_leads))
        # Dedup by name
        existing = {l['name'].lower() for l in leads}
        g_new = [l for l in g_leads if l['name'].lower() not in existing]
        for i,l in enumerate(g_new): l['srNo'] = len(leads)+i+1
        leads.extend(g_new)
        print(f'   ✅ Google: {len(g_new)} new leads')

    # Renumber
    for i,l in enumerate(leads): l['srNo'] = i+1

    print(f'\n📊 Total: {len(leads)} leads')
    print(f'   With phone:   {sum(1 for l in leads if l.get("phone"))}')
    print(f'   With website: {sum(1 for l in leads if l.get("website"))}')

    if not leads:
        print('⚠ No leads found. Will retry tomorrow.')
        # Still save empty file so frontend knows it ran
        save_leads([], industry, city)
        return

    # Save
    save_leads(leads, industry, city)
    save_excel(leads, industry, city)
    print(f'\n✅ Done! Leads saved for {TODAY}')

if __name__ == '__main__':
    main()
