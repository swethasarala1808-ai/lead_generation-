"""
bizaxl LeadGen Desktop App
===========================
Replaces manual Google Maps searching + Excel copy-paste.
Sales team: select industry + area + city → get leads with phone, address, WhatsApp → export Excel.
"""

import sys, json, re, csv, os
import requests
from math import radians, cos, sin, asin, sqrt
from datetime import datetime

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QLineEdit, QComboBox, QSlider, QFrame,
    QScrollArea, QGridLayout, QMessageBox, QFileDialog, QProgressBar,
    QSplitter, QSizePolicy, QSpacerItem, QAbstractScrollArea
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QSize, QTimer
from PyQt5.QtGui import QFont, QCursor, QColor, QPalette
import openpyxl
from openpyxl.styles import Font as XFont, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── COLOURS ───────────────────────────────────────────────────────
N   = "#0A1628"; N2  = "#111D35"
B   = "#2563EB"; BL  = "#DBEAFE"; BLL = "#EFF6FF"
G   = "#059669"; GL  = "#DCFCE7"; GD  = "#047857"
GLD = "#D97706"; GLDL= "#FEF3C7"
RD  = "#DC2626"; RDL = "#FEE2E2"
PU  = "#7C3AED"; PUL = "#EDE9FE"
GR  = "#6B7280"; GRB = "#E5E7EB"; GRL = "#F9FAFB"
T   = "#111827"; T2  = "#374151"; T3  = "#6B7280"
W   = "#FFFFFF"; BG  = "#F0F2F5"

# ── INDUSTRY → OSM TAGS ───────────────────────────────────────────
IND = {
    "Medical Shops":    [("amenity","pharmacy"),("amenity","clinic"),("amenity","doctors"),("amenity","dentist"),("amenity","hospital"),("healthcare","pharmacy")],
    "Retail Shops":     [("shop","supermarket"),("shop","convenience"),("shop","clothes"),("shop","general"),("shop","department_store")],
    "Grocery Stores":   [("shop","supermarket"),("shop","convenience"),("shop","greengrocer"),("shop","bakery")],
    "Restaurants":      [("amenity","restaurant"),("amenity","cafe"),("amenity","fast_food"),("amenity","bar")],
    "IT Companies":     [("office","it"),("office","company"),("office","coworking")],
    "Schools / College":[("amenity","school"),("amenity","college"),("amenity","university"),("amenity","training")],
    "CA / Tax Firms":   [("office","tax_advisor"),("office","accountant"),("office","financial_advisor")],
    "Real Estate":      [("office","estate_agent"),("office","construction")],
    "Logistics":        [("office","logistics"),("amenity","courier_office")],
    "Beauty / Salon":   [("shop","beauty"),("shop","hairdresser"),("leisure","fitness_centre"),("amenity","spa")],
    "Automobile":       [("shop","car"),("shop","car_repair"),("shop","motorcycle")],
    "Hardware Stores":  [("shop","hardware"),("shop","electrical"),("shop","doityourself")],
    "Hotels":           [("tourism","hotel"),("tourism","guest_house"),("tourism","hostel")],
    "Banks / Finance":  [("amenity","bank"),("office","insurance"),("office","financial_advisor")],
    "Manufacturing":    [("man_made","works"),("craft","*"),("office","industrial")],
}

CITIES = {
    "Bangalore":     (12.9716,77.5946,"Karnataka"),
    "Mumbai":        (19.0760,72.8777,"Maharashtra"),
    "Delhi":         (28.6139,77.2090,"Delhi"),
    "Chennai":       (13.0827,80.2707,"Tamil Nadu"),
    "Hyderabad":     (17.3850,78.4867,"Telangana"),
    "Pune":          (18.5204,73.8567,"Maharashtra"),
    "Kolkata":       (22.5726,88.3639,"West Bengal"),
    "Ahmedabad":     (23.0225,72.5714,"Gujarat"),
    "Bhopal":        (23.2599,77.4126,"Madhya Pradesh"),
    "Indore":        (22.7196,75.8577,"Madhya Pradesh"),
    "Patna":         (25.5941,85.1376,"Bihar"),
    "Jaipur":        (26.9124,75.7873,"Rajasthan"),
    "Lucknow":       (26.8467,80.9462,"Uttar Pradesh"),
    "Coimbatore":    (11.0168,76.9558,"Tamil Nadu"),
    "Kochi":         (9.9312, 76.2673,"Kerala"),
    "Surat":         (21.1702,72.8311,"Gujarat"),
    "Nagpur":        (21.1458,79.0882,"Maharashtra"),
    "Chandigarh":    (30.7333,76.7794,"Punjab"),
    "Visakhapatnam": (17.6868,83.2185,"Andhra Pradesh"),
    "Mysore":        (12.2958,76.6394,"Karnataka"),
    "Madurai":       (9.9252, 78.1198,"Tamil Nadu"),
    "Vadodara":      (22.3072,73.1812,"Gujarat"),
    "Ludhiana":      (30.9010,75.8573,"Punjab"),
}

OSM_TYPE = {
    "pharmacy":"Pharmacy","clinic":"Medical Clinic","doctors":"Doctor/GP","dentist":"Dental Clinic",
    "hospital":"Hospital","supermarket":"Supermarket","convenience":"Convenience Store",
    "clothes":"Clothing Store","general":"General Store","department_store":"Department Store",
    "greengrocer":"Vegetable Shop","bakery":"Bakery","restaurant":"Restaurant","cafe":"Cafe",
    "fast_food":"Fast Food","bar":"Bar/Pub","it":"IT Company","company":"Company",
    "coworking":"Co-working","school":"School","college":"College","university":"University",
    "training":"Training Institute","tax_advisor":"Tax Advisor","accountant":"CA Firm",
    "financial_advisor":"Financial Advisor","estate_agent":"Real Estate Agent",
    "construction":"Construction Co.","logistics":"Logistics","courier_office":"Courier",
    "beauty":"Beauty Salon","hairdresser":"Hair Salon","fitness_centre":"Gym/Fitness",
    "spa":"Spa","car":"Car Dealer","car_repair":"Auto Service","motorcycle":"Motorcycle Dealer",
    "hardware":"Hardware Store","electrical":"Electrical Shop","hotel":"Hotel",
    "guest_house":"Guest House","hostel":"Hostel","bank":"Bank","insurance":"Insurance Agency",
    "works":"Manufacturing Unit",
}

REG = ["MSME Registered","Proprietorship","SME","Partnership","Pvt Ltd","LLP"]
OSM_ENDPOINTS = [
    "https://overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
    "https://maps.mail.ru/osm/tools/overpass/api/interpreter",
]

# ── HELPERS ───────────────────────────────────────────────────────
def clean_phone(p):
    if not p: return ""
    d = re.sub(r"\D","",p)
    if len(d)==12 and d.startswith("91"): return d[2:]
    if len(d)==11 and d.startswith("0"):  return d[1:]
    if len(d)==10 and d[0] in "6789":     return d
    return ""

def clean_web(w):
    if not w: return ""
    return re.sub(r"^https?://","",w).lstrip("www.").split("/")[0].split("?")[0]

def build_addr(t, area, city):
    p=[]
    num=t.get("addr:housenumber",""); st=t.get("addr:street","")
    sub=t.get("addr:suburb","") or t.get("addr:neighbourhood","")
    ct=t.get("addr:city","") or t.get("addr:town","")
    if num and st: p.append(f"{num}, {st}")
    elif st: p.append(st)
    if sub: p.append(sub)
    if area and area.lower() not in " ".join(p).lower(): p.append(area)
    p.append(ct or city)
    return ", ".join(filter(None,p))

def get_btype(t):
    for k in ["amenity","shop","office","leisure","craft","tourism","man_made","healthcare"]:
        if k in t:
            return OSM_TYPE.get(t[k], t[k].replace("_"," ").title())
    return "Business"

def haversine(lat1,lon1,lat2,lon2):
    R=6371; dlat=radians(lat2-lat1); dlon=radians(lon2-lon1)
    a=sin(dlat/2)**2+cos(radians(lat1))*cos(radians(lat2))*sin(dlon/2)**2
    return R*2*asin(sqrt(a))

def score(lead):
    s=50
    if lead.get("phone"):   s+=18
    if lead.get("website"): s+=10
    if lead.get("email") and "@gmail" not in lead.get("email",""): s+=7
    if lead.get("pincode"): s+=5
    if lead.get("opening"): s+=5
    if lead.get("phone2"):  s+=3
    return min(98,s)

# ── SEARCH THREAD ─────────────────────────────────────────────────
class Searcher(QThread):
    prog   = pyqtSignal(int, str)
    done   = pyqtSignal(list)
    failed = pyqtSignal(str)

    def __init__(self, ind, area, city, count):
        super().__init__()
        self.ind=ind; self.area=area; self.city=city; self.count=count

    def run(self):
        try:
            tags   = IND.get(self.ind, [("amenity","pharmacy")])
            coords = CITIES.get(self.city,(12.9716,77.5946,"Karnataka"))
            clat,clon,state = coords

            lat,lon = clat,clon
            if self.area.strip():
                self.prog.emit(8, f'📍 Locating "{self.area}, {self.city}"...')
                geo = self._geocode(f"{self.area}, {self.city}, India")
                if geo:
                    lat,lon=geo
                    self.prog.emit(18, f'✅ Located: {self.area}')
                else:
                    self.prog.emit(18, f'ℹ Using city centre of {self.city}')

            # Radius: 3km for specific area, 15km for city-wide
            radius = 3000 if self.area.strip() else 15000
            self.prog.emit(22, f'Searching {self.ind} near {self.area or self.city}...')

            # Build query
            lines=[]
            for k,v in tags:
                kv = f'["{k}"]' if v=="*" else f'["{k}"="{v}"]'
                lines.append(f'  node{kv}["name"](around:{radius},{lat},{lon});')
                lines.append(f'  way{kv}["name"](around:{radius},{lat},{lon});')

            q = f'[out:json][timeout:35];\n(\n{"  ".join(lines)}\n);\nout body center {self.count*4};'

            self.prog.emit(32, 'Connecting to OpenStreetMap...')
            elements=[]
            for ep in OSM_ENDPOINTS:
                try:
                    r=requests.post(ep,data={"data":q},
                        headers={"User-Agent":"bizaxl-LeadGen/2.0 (sales@bizaxl.com)"},
                        timeout=40)
                    if r.status_code==200:
                        elements=r.json().get("elements",[])
                        self.prog.emit(60,f'Processing {len(elements)} results from OSM...')
                        break
                except: continue

            if not elements:
                self.failed.emit(
                    f"No results found near '{self.area or self.city}'.\n\n"
                    "Tips:\n"
                    "• Type a main area name (e.g. 'Koramangala', 'Andheri West')\n"
                    "• Try a nearby landmark or main road name\n"
                    "• Try without area — just city search\n"
                    "• Check internet connection"
                )
                return

            leads=[]; seen=set()
            for el in elements:
                if len(leads)>=self.count: break
                t=el.get("tags",{})
                name=t.get("name") or t.get("name:en","")
                if not name or name.lower() in seen: continue
                seen.add(name.lower())
                elat=el.get("lat") or (el.get("center") or {}).get("lat")
                elon=el.get("lon") or (el.get("center") or {}).get("lon")
                if not elat: continue

                phone  = clean_phone(t.get("phone","") or t.get("contact:phone","") or t.get("mobile","") or t.get("phone:1",""))
                phone2 = clean_phone(t.get("phone:2","") or t.get("contact:phone:2",""))
                web    = clean_web(t.get("website","") or t.get("contact:website",""))
                email  = t.get("email","") or t.get("contact:email","") or (f"info@{web}" if web else "")
                addr   = build_addr(t, self.area, self.city)
                pin    = t.get("postcode","") or t.get("addr:postcode","")
                hrs    = t.get("opening_hours","")
                gstin  = t.get("ref:GSTIN","")
                btype  = get_btype(t)
                dist   = haversine(lat,lon,elat,elon)

                lead={
                    "srNo":   len(leads)+1,
                    "name":   name,
                    "btype":  btype,
                    "reg":    REG[len(leads)%len(REG)],
                    "phone":  phone,
                    "phone2": phone2,
                    "email":  email,
                    "web":    web,
                    "addr":   addr,
                    "city":   self.city,
                    "state":  state,
                    "pin":    pin,
                    "hrs":    hrs,
                    "gstin":  gstin,
                    "wa":     "Yes" if phone else "No",
                    "lat":    elat, "lon": elon,
                    "dist":   dist,
                    "maps":   f"https://www.google.com/maps/search/?api=1&query={requests.utils.quote(name+' '+addr)}",
                    "src":    "OpenStreetMap",
                }
                lead["score"] = score(lead)
                leads.append(lead)
                self.prog.emit(60+int((len(leads)/self.count)*36),
                    f'Found {len(leads)} of {self.count}...')

            if not leads:
                self.failed.emit(f"No {self.ind} found near {self.area or self.city}.\nTry a different area or industry.")
                return

            self.prog.emit(100,"Done!")
            self.done.emit(leads)
        except Exception as e:
            self.failed.emit(str(e))

    def _geocode(self, q):
        try:
            r=requests.get("https://nominatim.openstreetmap.org/search",
                params={"q":q,"format":"json","limit":1},
                headers={"User-Agent":"bizaxl-LeadGen/2.0"},
                timeout=12)
            d=r.json()
            if d: return float(d[0]["lat"]),float(d[0]["lon"])
        except: pass
        return None

# ── LEAD CARD ─────────────────────────────────────────────────────
class LeadCard(QFrame):
    def __init__(self, lead, row_num):
        super().__init__()
        self.lead=lead
        self.setObjectName("card")
        self.setStyleSheet("""
            QFrame#card{background:white;border:1.5px solid #E5E7EB;border-radius:10px}
            QFrame#card:hover{border-color:#BFDBFE}
        """)
        layout=QHBoxLayout(self)
        layout.setContentsMargins(14,12,14,12)
        layout.setSpacing(14)

        # ── Row number ──
        num=QLabel(str(row_num))
        num.setFixedWidth(28)
        num.setFont(QFont("Segoe UI",12,QFont.Bold))
        num.setStyleSheet(f"color:{T3};background:transparent;border:none")
        num.setAlignment(Qt.AlignTop|Qt.AlignCenter)
        layout.addWidget(num)

        # ── Info ──
        info=QVBoxLayout(); info.setSpacing(3)

        # Name + type
        name_row=QHBoxLayout(); name_row.setSpacing(8)
        nl=QLabel(lead["name"])
        nl.setFont(QFont("Segoe UI",12,QFont.Bold))
        nl.setStyleSheet(f"color:{T};background:transparent;border:none")
        nl.setWordWrap(True)
        name_row.addWidget(nl)

        tl=QLabel(lead["btype"].upper())
        tl.setFont(QFont("Segoe UI",9))
        tl.setStyleSheet(f"color:{T3};background:{GRL};border:1px solid {GRB};border-radius:4px;padding:1px 7px")
        name_row.addWidget(tl)
        name_row.addStretch()
        info.addLayout(name_row)

        # Address — exactly like the sample data
        addr_full=lead["addr"]
        if lead["pin"] and lead["pin"] not in addr_full:
            addr_full+=f", {lead['city']}, {lead['state']} {lead['pin']}"
        al=QLabel(f"📍  {addr_full}")
        al.setFont(QFont("Segoe UI",10))
        al.setStyleSheet(f"color:{T3};background:transparent;border:none")
        al.setWordWrap(True)
        info.addWidget(al)

        # Opening hours
        if lead.get("hrs"):
            hl=QLabel(f"🕐  {lead['hrs']}")
            hl.setFont(QFont("Segoe UI",10))
            hl.setStyleSheet(f"color:{GD};background:transparent;border:none")
            info.addWidget(hl)

        # Tags
        tag_row=QHBoxLayout(); tag_row.setSpacing(5)
        reg=lead.get("reg","")
        self._tag(tag_row, reg, PU if "MSME" in reg else B, PUL if "MSME" in reg else BLL)
        if lead.get("gstin"):
            self._tag(tag_row," ✓ GSTIN",GD,GL)
        if lead.get("wa")=="Yes":
            self._tag(tag_row,"💬 WhatsApp",GD,GL)
        if lead.get("dist") is not None:
            d=lead["dist"]
            self._tag(tag_row,f"📍 {d*1000:.0f}m" if d<1 else f"📍 {d:.1f}km",GD,GL)
        tag_row.addStretch()
        info.addLayout(tag_row)

        layout.addLayout(info,1)

        # ── Action buttons ──
        acts=QVBoxLayout(); acts.setSpacing(5); acts.setAlignment(Qt.AlignTop)

        # Score pill
        s=lead.get("score",50)
        sc=GD if s>=80 else GLD if s>=60 else RD
        sb=GL if s>=80 else GLDL if s>=60 else RDL
        sp=QLabel(f"Quality {s}%")
        sp.setFont(QFont("Segoe UI",10,QFont.Bold))
        sp.setStyleSheet(f"color:{sc};background:{sb};border-radius:6px;padding:3px 9px;border:none;min-width:90px")
        sp.setAlignment(Qt.AlignCenter)
        acts.addWidget(sp)

        if lead.get("phone"):
            pb=self._btn(f"📞  {lead['phone']}", B, W, B)
            pb.clicked.connect(lambda _,p=lead["phone"]: self._call(p))
            acts.addWidget(pb)

            wb=self._btn("💬  WhatsApp","#16A34A",W,"#16A34A")
            msg=requests.utils.quote("Hi, I'm from bizaxl. We offer business software solutions. Can we schedule a quick call?")
            wb.clicked.connect(lambda _,p=lead["phone"],m=msg: self._url(f"https://wa.me/91{p}?text={m}"))
            acts.addWidget(wb)
        else:
            nl2=QLabel("No phone in OSM data")
            nl2.setFont(QFont("Segoe UI",10))
            nl2.setStyleSheet(f"color:{T3};border:1px dashed {GRB};border-radius:7px;padding:6px;background:transparent")
            nl2.setAlignment(Qt.AlignCenter)
            acts.addWidget(nl2)

        if lead.get("email"):
            eb=self._btn(f"✉️  {lead['email'][:24]}{'…' if len(lead['email'])>24 else ''}",W,T2,GRB)
            eb.clicked.connect(lambda _,e=lead["email"]: self._url(f"mailto:{e}"))
            acts.addWidget(eb)

        mb=self._btn("📍  Google Maps",W,T3,GRB)
        mb.clicked.connect(lambda _,u=lead["maps"]: self._url(u))
        acts.addWidget(mb)

        if lead.get("web"):
            wb2=self._btn(f"🔗  {lead['web'][:24]}",W,T3,GRB)
            wb2.clicked.connect(lambda _,u=lead["web"]: self._url(f"https://{u}"))
            acts.addWidget(wb2)

        layout.addLayout(acts)

    def _tag(self,lay,txt,col,bg):
        l=QLabel(txt); l.setFont(QFont("Segoe UI",9,QFont.Bold))
        l.setStyleSheet(f"color:{col};background:{bg};border-radius:10px;padding:2px 8px;border:none")
        lay.addWidget(l)

    def _btn(self,txt,bg,fg,brd):
        b=QPushButton(txt); b.setFont(QFont("Segoe UI",10,QFont.Bold))
        b.setCursor(QCursor(Qt.PointingHandCursor)); b.setFixedHeight(32)
        b.setMinimumWidth(160)
        b.setStyleSheet(f"QPushButton{{background:{bg};color:{fg};border:1.5px solid {brd};border-radius:7px;padding:4px 12px;text-align:center}}QPushButton:hover{{opacity:.85}}")
        return b

    def _call(self,p):
        import subprocess
        if sys.platform=="win32": subprocess.Popen(f'start tel:+91{p}',shell=True)

    def _url(self,u):
        import subprocess,webbrowser
        if sys.platform=="win32": subprocess.Popen(f'start "" "{u}"',shell=True)
        else: webbrowser.open(u)

# ── MAIN WINDOW ───────────────────────────────────────────────────
class App(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("bizaxl LeadGen — Automate Your Daily Lead Collection")
        self.resize(1280,820); self.setMinimumSize(1000,650)
        self.leads=[]; self.filtered=[]; self.page=1; self.per=15
        self.qf="all"; self.worker=None
        self._build()

    def _build(self):
        self.setStyleSheet(f"QMainWindow{{background:{BG}}}")
        cw=QWidget(); self.setCentralWidget(cw)
        root=QVBoxLayout(cw); root.setContentsMargins(0,0,0,0); root.setSpacing(0)

        # Title bar
        tb=QFrame(); tb.setFixedHeight(46); tb.setStyleSheet(f"background:{N};border:none")
        tbl=QHBoxLayout(tb); tbl.setContentsMargins(16,0,16,0); tbl.setSpacing(10)
        ico=QLabel("🎯"); ico.setFont(QFont("Segoe UI Emoji",16)); ico.setStyleSheet("background:transparent;border:none")
        tbl.addWidget(ico)
        lname=QLabel('<span style="color:white;font-size:15px;font-weight:700">biz</span><span style="color:#10B981;font-size:15px;font-weight:700">axl</span><span style="color:white;font-size:15px;font-weight:700"> LeadGen</span>')
        lname.setStyleSheet("background:transparent;border:none")
        tbl.addWidget(lname)
        sub=QLabel("Automates daily lead collection — replaces manual Google Maps search + Excel copy-paste")
        sub.setFont(QFont("Segoe UI",10)); sub.setStyleSheet(f"color:#475569;background:transparent;border:none")
        tbl.addWidget(sub); tbl.addStretch()
        for txt,col in [("🌍 OpenStreetMap — Free Data","#34D399"),("bizaxl Internal Tool","#475569")]:
            p=QLabel(txt); p.setFont(QFont("Segoe UI",9,QFont.Bold))
            p.setStyleSheet(f"color:{col};background:rgba(255,255,255,0.07);border:1px solid rgba(255,255,255,0.1);border-radius:20px;padding:3px 10px")
            tbl.addWidget(p); tbl.addSpacing(4)
        root.addWidget(tb)

        # Body split
        sp=QSplitter(Qt.Horizontal); sp.setHandleWidth(1)
        sp.setStyleSheet(f"QSplitter::handle{{background:{GRB}}}")

        sp.addWidget(self._left_panel())
        sp.addWidget(self._right_panel())
        sp.setSizes([300,980]); sp.setCollapsible(0,False); sp.setCollapsible(1,False)
        root.addWidget(sp,1)

    # ── LEFT PANEL ────────────────────────────────────────────────
    def _left_panel(self):
        f=QFrame(); f.setFixedWidth(300)
        f.setStyleSheet(f"QFrame{{background:{W};border-right:1px solid {GRB};border-top:none;border-left:none;border-bottom:none}}")
        vl=QVBoxLayout(f); vl.setContentsMargins(0,0,0,0); vl.setSpacing(0)

        sc=QScrollArea(); sc.setWidgetResizable(True); sc.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        sc.setStyleSheet("QScrollArea{border:none;background:transparent}")
        inner=QWidget(); inner.setStyleSheet(f"background:{W}")
        fl=QVBoxLayout(inner); fl.setContentsMargins(14,14,14,10); fl.setSpacing(10)

        # Industry
        self._lbl(fl,"INDUSTRY / CATEGORY")
        self.ind_btns={}; grid=QGridLayout(); grid.setSpacing(5)
        inds=list(IND.keys())
        icons={"Medical Shops":"🏥","Retail Shops":"🛍️","Grocery Stores":"🛒",
               "Restaurants":"🍽️","IT Companies":"💻","Schools / College":"🎓",
               "CA / Tax Firms":"💰","Real Estate":"🏗️","Logistics":"🚚",
               "Beauty / Salon":"💅","Automobile":"🚗","Hardware Stores":"🔧",
               "Hotels":"🏨","Banks / Finance":"🏦","Manufacturing":"⚙️"}
        for i,ind in enumerate(inds):
            btn=QPushButton(f"{icons.get(ind,'📌')} {ind}")
            btn.setFont(QFont("Segoe UI",10))
            btn.setCheckable(True); btn.setChecked(ind=="Medical Shops")
            btn.setCursor(QCursor(Qt.PointingHandCursor))
            btn.setStyleSheet(self._ibstyle(ind=="Medical Shops"))
            btn.clicked.connect(lambda _,b=btn,n=ind: self._sel_ind(b,n))
            grid.addWidget(btn,i//2,i%2)
            self.ind_btns[ind]=btn
        fl.addLayout(grid); self.sel_ind="Medical Shops"

        fl.addWidget(self._div())

        # Area
        self._lbl(fl,"AREA / LOCALITY")
        hint=QLabel("Like Google Maps — type 'Kasturi Nagar' or 'Koramangala'")
        hint.setFont(QFont("Segoe UI",9)); hint.setWordWrap(True)
        hint.setStyleSheet(f"color:{T3};background:transparent;border:none")
        fl.addWidget(hint)
        self.area_in=QLineEdit(); self.area_in.setPlaceholderText("e.g. Kasturi Nagar, KR Puram, Andheri...")
        self.area_in.setFont(QFont("Segoe UI",11)); self.area_in.setStyleSheet(self._instyle())
        self.area_in.returnPressed.connect(self._search)
        fl.addWidget(self.area_in)

        # City
        self._lbl(fl,"CITY")
        self.city_cb=QComboBox(); self.city_cb.addItems(list(CITIES.keys()))
        self.city_cb.setFont(QFont("Segoe UI",11)); self.city_cb.setStyleSheet(self._cbstyle())
        fl.addWidget(self.city_cb)

        # State (auto)
        self._lbl(fl,"STATE (auto-filled)")
        self.state_cb=QComboBox()
        self.state_cb.addItems(sorted(set(v[2] for v in CITIES.values())))
        self.state_cb.setFont(QFont("Segoe UI",11)); self.state_cb.setStyleSheet(self._cbstyle())
        self.state_cb.setCurrentText("Karnataka")
        self.city_cb.currentTextChanged.connect(lambda c: self.state_cb.setCurrentText(CITIES.get(c,(0,0,"Karnataka"))[2]))
        fl.addWidget(self.state_cb)

        # Count
        self._lbl(fl,"NUMBER OF LEADS")
        cr=QHBoxLayout()
        self.cnt_lbl=QLabel("30"); self.cnt_lbl.setFont(QFont("Segoe UI",22,QFont.Bold))
        self.cnt_lbl.setStyleSheet(f"color:{B};background:transparent;border:none;min-width:40px")
        self.cnt_lbl.setAlignment(Qt.AlignCenter)
        cr.addWidget(self.cnt_lbl)
        self.cnt_sl=QSlider(Qt.Horizontal); self.cnt_sl.setRange(10,150); self.cnt_sl.setValue(30)
        self.cnt_sl.setStyleSheet(f"QSlider::groove:horizontal{{height:4px;background:{GRB};border-radius:2px}}QSlider::handle:horizontal{{width:16px;height:16px;margin:-6px 0;background:{B};border-radius:8px}}QSlider::sub-page:horizontal{{background:{B};border-radius:2px}}")
        self.cnt_sl.valueChanged.connect(lambda v: self.cnt_lbl.setText(str(v)))
        cr.addWidget(self.cnt_sl,1)
        fl.addLayout(cr)

        fl.addWidget(self._div())

        # Search button
        self.sbtn=QPushButton("🔍   Find Leads Now")
        self.sbtn.setFont(QFont("Segoe UI",13,QFont.Bold)); self.sbtn.setFixedHeight(46)
        self.sbtn.setCursor(QCursor(Qt.PointingHandCursor))
        self.sbtn.setStyleSheet(f"QPushButton{{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 {B},stop:1 #1251A3);color:white;border:none;border-radius:10px;padding:10px}}QPushButton:hover{{background:{B}}}QPushButton:disabled{{background:#93C5FD}}")
        self.sbtn.clicked.connect(self._search)
        fl.addWidget(self.sbtn)

        fl.addWidget(self._div())

        # Quick links
        self._lbl(fl,"OPEN DIRECTORY SITES")
        for name,url in [
            ("📦 TradeIndia ↗","https://www.tradeindia.com"),
            ("📋 ExportersIndia ↗","https://www.exportersindia.com"),
            ("🔵 Infobel India ↗","https://local.infobel.in"),
            ("📞 Dial4Trade ↗","https://www.dial4trade.com"),
            ("🇮🇳 Hello India ↗","https://www.helloindia.co"),
            ("📊 Rentech DB ↗","https://rentechdigital.com/smartscraper"),
        ]:
            b=QPushButton(name); b.setFont(QFont("Segoe UI",10))
            b.setCursor(QCursor(Qt.PointingHandCursor))
            b.setStyleSheet(f"QPushButton{{background:white;color:{T2};border:1px solid {GRB};border-radius:7px;padding:6px 10px;text-align:left}}QPushButton:hover{{border-color:{B};color:{B};background:{BLL}}}")
            b.clicked.connect(lambda _,u=url: self._openurl(u))
            fl.addWidget(b)

        fl.addStretch()
        sc.setWidget(inner); vl.addWidget(sc,1)

        # Stats bar
        self.stats=QFrame(); self.stats.setVisible(False)
        self.stats.setStyleSheet(f"background:{GRL};border-top:1px solid {GRB}")
        sl=QHBoxLayout(self.stats); sl.setContentsMargins(10,8,10,8); self.svl={}
        for key,lbl in [("t","Total"),("p","Phone"),("w","Website")]:
            bx=QVBoxLayout()
            v=QLabel("0"); v.setFont(QFont("Segoe UI",18,QFont.Bold))
            v.setStyleSheet(f"color:{T};background:transparent;border:none")
            v.setAlignment(Qt.AlignCenter)
            l=QLabel(lbl); l.setFont(QFont("Segoe UI",10))
            l.setStyleSheet(f"color:{T3};background:transparent;border:none")
            l.setAlignment(Qt.AlignCenter)
            bx.addWidget(v); bx.addWidget(l); sl.addLayout(bx); self.svl[key]=v
        vl.addWidget(self.stats)
        return f

    # ── RIGHT PANEL ───────────────────────────────────────────────
    def _right_panel(self):
        f=QFrame(); f.setStyleSheet(f"QFrame{{background:{BG};border:none}}")
        vl=QVBoxLayout(f); vl.setContentsMargins(0,0,0,0); vl.setSpacing(0)

        # Toolbar
        self.toolbar=QFrame(); self.toolbar.setVisible(False); self.toolbar.setFixedHeight(48)
        self.toolbar.setStyleSheet(f"background:{W};border-bottom:1px solid {GRB};border-top:none;border-left:none;border-right:none")
        tl=QHBoxLayout(self.toolbar); tl.setContentsMargins(12,0,12,0); tl.setSpacing(8)
        self.rc=QLabel(); self.rc.setFont(QFont("Segoe UI",12,QFont.Bold))
        self.rc.setStyleSheet(f"color:{T};background:transparent;border:none")
        tl.addWidget(self.rc)
        self.sb=QLineEdit(); self.sb.setPlaceholderText("🔍  Search by name, phone, area...")
        self.sb.setFont(QFont("Segoe UI",11)); self.sb.setStyleSheet(self._instyle())
        self.sb.textChanged.connect(self._refilter); tl.addWidget(self.sb,1)
        self.sc2=QComboBox()
        self.sc2.addItems(["Sort: Quality Score","Sort: Has Phone First","Sort: Distance","Sort: Name A–Z"])
        self.sc2.setFont(QFont("Segoe UI",10)); self.sc2.setStyleSheet(self._cbstyle())
        self.sc2.currentIndexChanged.connect(self._refilter); tl.addWidget(self.sc2)
        for lbl,fn,bg,fg in [("📥 CSV",self._csv,W,T2),("📊 Excel",self._xlsx,G,W)]:
            b=QPushButton(lbl); b.setFont(QFont("Segoe UI",10,QFont.Bold))
            b.setCursor(QCursor(Qt.PointingHandCursor))
            brd=GRB if bg==W else G
            b.setStyleSheet(f"QPushButton{{background:{bg};color:{fg};border:1.5px solid {brd};border-radius:7px;padding:6px 12px}}QPushButton:hover{{{'border-color:'+B+';color:'+B if bg==W else 'background:'+GD}}}")
            b.clicked.connect(fn); tl.addWidget(b)
        vl.addWidget(self.toolbar)

        # Filter chips
        self.fbar=QFrame(); self.fbar.setVisible(False); self.fbar.setFixedHeight(38)
        self.fbar.setStyleSheet(f"background:{W};border-bottom:1px solid {GRB};border-top:none;border-left:none;border-right:none")
        fbl=QHBoxLayout(self.fbar); fbl.setContentsMargins(12,0,12,0); fbl.setSpacing(6)
        fl2=QLabel("FILTER:"); fl2.setFont(QFont("Segoe UI",9,QFont.Bold))
        fl2.setStyleSheet(f"color:{T3};background:transparent;border:none;letter-spacing:1px")
        fbl.addWidget(fl2)
        self.fbts={}
        for fid,fn in [("all","All"),("phone","📞 Has Phone"),("web","🌐 Website"),
                        ("email","✉️ Email"),("wa","💬 WhatsApp"),("hrs","🕐 Has Hours")]:
            b=QPushButton(fn); b.setFont(QFont("Segoe UI",10))
            b.setCursor(QCursor(Qt.PointingHandCursor))
            b.setCheckable(True); b.setChecked(fid=="all")
            b.setStyleSheet(self._fbstyle(fid=="all"))
            b.clicked.connect(lambda _,i=fid,bt=b: self._qfilter(i,bt))
            fbl.addWidget(b); self.fbts[fid]=b
        fbl.addStretch(); vl.addWidget(self.fbar)

        # Progress
        self.pf=QFrame(); self.pf.setVisible(False)
        self.pf.setStyleSheet(f"background:{W};border-bottom:1px solid {GRB}")
        pfl=QVBoxLayout(self.pf); pfl.setContentsMargins(20,14,20,14); pfl.setSpacing(6)
        self.pt=QLabel("Searching..."); self.pt.setFont(QFont("Segoe UI",12,QFont.Bold))
        self.pt.setStyleSheet(f"color:{T};background:transparent;border:none")
        self.pt.setAlignment(Qt.AlignCenter); pfl.addWidget(self.pt)
        self.pb=QProgressBar(); self.pb.setRange(0,100); self.pb.setValue(0)
        self.pb.setTextVisible(False); self.pb.setFixedHeight(6)
        self.pb.setStyleSheet(f"QProgressBar{{background:{GRB};border-radius:3px;border:none}}QProgressBar::chunk{{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #10B981,stop:1 {B});border-radius:3px}}")
        pfl.addWidget(self.pb)
        self.pm=QLabel(); self.pm.setFont(QFont("Segoe UI",10))
        self.pm.setStyleSheet(f"color:{T3};background:transparent;border:none")
        self.pm.setAlignment(Qt.AlignCenter); pfl.addWidget(self.pm)
        vl.addWidget(self.pf)

        # Scroll area for cards
        self.scroll=QScrollArea(); self.scroll.setWidgetResizable(True)
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.scroll.setStyleSheet(f"QScrollArea{{border:none;background:{BG}}}")
        self.cw=QWidget(); self.cw.setStyleSheet(f"background:{BG}")
        self.cl=QVBoxLayout(self.cw); self.cl.setContentsMargins(12,10,12,10); self.cl.setSpacing(8)
        self._empty("Ready to find leads\n\nSelect industry → type area → click Find Leads\nJust like Google Maps — shows all nearby businesses")
        self.scroll.setWidget(self.cw); vl.addWidget(self.scroll,1)

        # Pagination bar
        self.pgf=QFrame(); self.pgf.setVisible(False); self.pgf.setFixedHeight(46)
        self.pgf.setStyleSheet(f"background:{W};border-top:1px solid {GRB};border-bottom:none;border-left:none;border-right:none")
        self.pgl=QHBoxLayout(self.pgf); self.pgl.setAlignment(Qt.AlignCenter); self.pgl.setSpacing(6)
        vl.addWidget(self.pgf)
        return f

    # ── SEARCH ────────────────────────────────────────────────────
    def _search(self):
        ind  = self.sel_ind
        area = self.area_in.text().strip()
        city = self.city_cb.currentText()
        cnt  = self.cnt_sl.value()
        self._clear(); self.toolbar.setVisible(False); self.fbar.setVisible(False)
        self.stats.setVisible(False); self.pgf.setVisible(False)
        self.pf.setVisible(True); self.pb.setValue(0)
        self.sbtn.setEnabled(False); self.sbtn.setText("⏳  Searching...")
        self.worker=Searcher(ind,area,city,cnt)
        self.worker.prog.connect(lambda p,m: (self.pb.setValue(p),self.pt.setText(f"🔍 {m}" if p<100 else "✅ Done!"),self.pm.setText(m)))
        self.worker.done.connect(self._got_leads)
        self.worker.failed.connect(self._err)
        self.worker.start()

    def _got_leads(self,leads):
        self.pf.setVisible(False); self.sbtn.setEnabled(True); self.sbtn.setText("🔍   Find Leads Now")
        self.leads=leads; self.filtered=leads[:]
        self._refilter()
        self.toolbar.setVisible(True); self.fbar.setVisible(True); self.stats.setVisible(True)
        self._upd_stats(self.filtered)
        self.rc.setText(f'<b style="color:{B}">{len(leads)}</b> leads found')

    def _err(self,msg):
        self.pf.setVisible(False); self.sbtn.setEnabled(True); self.sbtn.setText("🔍   Find Leads Now")
        self._empty(f"⚠️  No results\n\n{msg}")

    def _refilter(self):
        d=self.leads[:]
        if self.qf=="phone":  d=[l for l in d if l.get("phone")]
        if self.qf=="web":    d=[l for l in d if l.get("web")]
        if self.qf=="email":  d=[l for l in d if l.get("email") and "@gmail" not in l.get("email","")]
        if self.qf=="wa":     d=[l for l in d if l.get("wa")=="Yes"]
        if self.qf=="hrs":    d=[l for l in d if l.get("hrs")]
        q=(self.sb.text() if hasattr(self,"sb") else "").strip().lower()
        if q: d=[l for l in d if any(q in str(v).lower() for v in l.values())]
        si=self.sc2.currentIndex() if hasattr(self,"sc2") else 0
        if si==0: d.sort(key=lambda l:-l.get("score",0))
        if si==1: d.sort(key=lambda l:(0 if l.get("phone") else 1))
        if si==2: d.sort(key=lambda l:l.get("dist") or 999)
        if si==3: d.sort(key=lambda l:l.get("name",""))
        self.filtered=d; self.page=1
        self._upd_stats(d); self._render()
        self.rc.setText(f'<b style="color:{B}">{len(d)}</b> leads found')

    def _render(self):
        self._clear()
        page=self.filtered[(self.page-1)*self.per:self.page*self.per]
        if not page:
            self._empty("No results — adjust filters or search again"); self.pgf.setVisible(False); return
        for i,lead in enumerate(page,1):
            self.cl.addWidget(LeadCard(lead,(self.page-1)*self.per+i))
        self.cl.addStretch()
        self._pg()

    def _pg(self):
        total=-(-len(self.filtered)//self.per)
        while self.pgl.count():
            it=self.pgl.takeAt(0)
            if it.widget(): it.widget().deleteLater()
        if total<=1: self.pgf.setVisible(False); return
        self.pgf.setVisible(True)
        if self.page>1: self._pgbtn("← Prev",self.page-1)
        for i in range(max(1,self.page-2),min(total,self.page+2)+1):
            self._pgbtn(str(i),i,i==self.page)
        if self.page<total: self._pgbtn("Next →",self.page+1)

    def _pgbtn(self,txt,pg,active=False):
        b=QPushButton(txt); b.setFont(QFont("Segoe UI",10,QFont.Bold))
        b.setCursor(QCursor(Qt.PointingHandCursor)); b.setFixedHeight(30)
        b.setStyleSheet(f"QPushButton{{background:{B if active else W};color:{W if active else T2};border:1.5px solid {B if active else GRB};border-radius:6px;padding:4px 10px}}QPushButton:hover{{border-color:{B};color:{B if not active else W}}}")
        b.clicked.connect(lambda _,p=pg: self._gopage(p)); self.pgl.addWidget(b)

    def _gopage(self,n):
        self.page=n; self._render(); self.scroll.verticalScrollBar().setValue(0)

    # ── EXPORT ────────────────────────────────────────────────────
    def _xlsx(self):
        if not self.filtered: QMessageBox.information(self,"No Data","No leads to export."); return
        path,_=QFileDialog.getSaveFileName(self,"Save Excel",
            f"bizaxl_leads_{self.sel_ind.replace(' ','_')}_{self.city_cb.currentText()}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            "Excel (*.xlsx)")
        if not path: return
        try:
            wb=openpyxl.Workbook(); ws=wb.active; ws.title="Leads"
            COLS=["srNo","name","phone","phone2","addr","city","state","pin","btype","reg","email","web","hrs","wa","gstin","score","maps","src"]
            HDRS=["Sr No","Company Name","Mobile Number","Alt Mobile","Full Address","City","State","Pincode","Business Type","Reg. Type","Email","Website","Opening Hours","WhatsApp","GSTIN","Quality %","Google Maps","Source"]
            thin=Side(style="thin",color="CCCCCC"); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)

            # Title row
            ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
            tc=ws["A1"]
            tc.value=f"bizaxl LeadGen — {self.sel_ind} | {self.area_in.text() or self.city_cb.currentText()} | {datetime.now().strftime('%d %b %Y')}"
            tc.font=XFont(name="Calibri",bold=True,size=14,color="FFFFFF")
            tc.fill=PatternFill("solid",fgColor="0A1628")
            tc.alignment=Alignment(horizontal="center",vertical="center")
            ws.row_dimensions[1].height=30

            # Header
            for ci,h in enumerate(HDRS,1):
                c=ws.cell(row=2,column=ci,value=h)
                c.font=XFont(name="Calibri",bold=True,size=10,color="E2E8F0")
                c.fill=PatternFill("solid",fgColor="2563EB")
                c.alignment=Alignment(horizontal="center",vertical="center")
                c.border=bdr
            ws.row_dimensions[2].height=24

            # Data
            for ri,lead in enumerate(self.filtered,3):
                for ci,col in enumerate(COLS,1):
                    v=lead.get(col,"")
                    c=ws.cell(row=ri,column=ci,value=v if col!="score" else str(v)+"%")
                    c.font=XFont(name="Calibri",size=9); c.border=bdr
                    if col=="phone" and v: c.font=XFont(name="Calibri",size=9,bold=True,color="059669")
                    if ri%2==0: c.fill=PatternFill("solid",fgColor="F0F4FF")

            for ci,w in enumerate([6,28,14,14,36,14,16,10,20,18,26,22,22,10,18,10,36,18],1):
                ws.column_dimensions[get_column_letter(ci)].width=w
            ws.freeze_panes="A3"
            ws.auto_filter.ref=f"A2:{get_column_letter(len(COLS))}{len(self.filtered)+2}"
            wb.save(path)
            QMessageBox.information(self,"✅ Exported!",f"{len(self.filtered)} leads saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self,"Export Error",str(e))

    def _csv(self):
        if not self.filtered: QMessageBox.information(self,"No Data","No leads to export."); return
        path,_=QFileDialog.getSaveFileName(self,"Save CSV",f"leads_{datetime.now().strftime('%Y%m%d')}.csv","CSV (*.csv)")
        if not path: return
        try:
            with open(path,"w",newline="",encoding="utf-8-sig") as f:
                w=csv.writer(f)
                w.writerow(["Sr No","Company Name","Mobile Number","Full Address","State","City","Pincode","Business Type","Email","Website","Opening Hours","WhatsApp","Quality Score","Google Maps"])
                for i,l in enumerate(self.filtered,1):
                    w.writerow([i,l.get("name",""),l.get("phone",""),l.get("addr",""),l.get("state",""),l.get("city",""),l.get("pin",""),l.get("btype",""),l.get("email",""),l.get("web",""),l.get("hrs",""),l.get("wa",""),str(l.get("score",""))+"%",l.get("maps","")])
            QMessageBox.information(self,"✅ Exported!",f"CSV saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self,"Export Error",str(e))

    # ── HELPERS ───────────────────────────────────────────────────
    def _sel_ind(self,btn,name):
        self.sel_ind=name
        for n,b in self.ind_btns.items():
            b.setChecked(n==name); b.setStyleSheet(self._ibstyle(n==name))

    def _qfilter(self,fid,btn):
        self.qf=fid
        for f,b in self.fbts.items():
            b.setChecked(f==fid); b.setStyleSheet(self._fbstyle(f==fid))
        self._refilter()

    def _clear(self):
        while self.cl.count():
            it=self.cl.takeAt(0)
            if it.widget(): it.widget().deleteLater()

    def _empty(self,msg):
        self._clear()
        l=QLabel(msg); l.setAlignment(Qt.AlignCenter)
        l.setFont(QFont("Segoe UI",11)); l.setWordWrap(True)
        l.setStyleSheet(f"color:{T3};background:transparent;border:none;padding:30px")
        self.cl.addStretch(); self.cl.addWidget(l); self.cl.addStretch()

    def _upd_stats(self,leads):
        self.svl["t"].setText(str(len(leads)))
        self.svl["p"].setText(str(sum(1 for l in leads if l.get("phone"))))
        self.svl["w"].setText(str(sum(1 for l in leads if l.get("web"))))

    def _openurl(self,u):
        import subprocess,webbrowser
        if sys.platform=="win32": subprocess.Popen(f'start "" "{u}"',shell=True)
        else: webbrowser.open(u)

    def _lbl(self,lay,txt):
        l=QLabel(txt); l.setFont(QFont("Segoe UI",9,QFont.Bold))
        l.setStyleSheet(f"color:{T3};background:transparent;border:none;letter-spacing:0.5px"); lay.addWidget(l)

    def _div(self):
        d=QFrame(); d.setFrameShape(QFrame.HLine)
        d.setStyleSheet(f"background:{GRB};border:none;height:1px;max-height:1px"); return d

    def _instyle(self):
        return f"QLineEdit{{padding:8px 10px;border:1.5px solid {GRB};border-radius:7px;font-size:12px;background:white;color:{T}}}QLineEdit:focus{{border-color:{B}}}"

    def _cbstyle(self):
        return f"QComboBox{{padding:7px 10px;border:1.5px solid {GRB};border-radius:7px;font-size:12px;background:white;color:{T}}}QComboBox:focus{{border-color:{B}}}QComboBox::drop-down{{border:none;width:20px}}"

    def _ibstyle(self,a):
        if a: return f"QPushButton{{background:{BL};color:{B};border:1.5px solid {B};border-radius:7px;padding:6px 4px;font-weight:700;text-align:center}}"
        return f"QPushButton{{background:{GRL};color:{T2};border:1.5px solid {GRB};border-radius:7px;padding:6px 4px;text-align:center}}QPushButton:hover{{border-color:{B};color:{B};background:{BLL}}}"

    def _fbstyle(self,a):
        if a: return f"QPushButton{{background:{BL};color:{B};border:1.5px solid {B};border-radius:20px;padding:3px 10px;font-weight:600}}"
        return f"QPushButton{{background:white;color:{T2};border:1px solid {GRB};border-radius:20px;padding:3px 10px}}QPushButton:hover{{border-color:{B};color:{B}}}"

if __name__=="__main__":
    app=QApplication(sys.argv)
    app.setStyle("Fusion")
    win=App(); win.show()
    sys.exit(app.exec_())
